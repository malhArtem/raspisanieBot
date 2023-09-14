import openpyxl

from func import parse_kurs, cell_value, db_start




def parse_xl():
    book = openpyxl.open("xl.xlsx")
    sheet = book.active
    print(sheet.max_row)
    print(sheet.max_column)
    col_begin = 0

    parse_kurs(col_begin)
    for j in range(sheet.max_column):
        if (cell_value(sheet, 4, j)) is None and (cell_value(sheet, 4, j + 1)) is not None:
            j += 1
            col_begin = j
            parse_kurs(col_begin)


db_start()
parse_xl()
