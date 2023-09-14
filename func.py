from aiogram import types
import sqlite3 as sq
import openpyxl
from openpyxl.cell.cell import MergedCell

from config import *
from main import bot


def db_start():
    global db, cur
    db = sq.connect('rasp.db')    # подключение к базе данных
    cur = db.cursor()             # создание курсора для взаимодействия с базой данных
    create_table_users()


def create_table_rasp(kurs):

    cur.execute("""CREATE TABLE IF NOT EXISTS '{}' 
                      (day TEXT, time TEXT,specialty TEXT, groupp TEXT, subj1 TEXT, 
                      teach1 TEXT, clas1 TEXT, subj2 TEXT, teach3 TEXT, clas2 TEXT)""".format(kurs))

    db.commit()


def create_table_users():

    cur.execute("""CREATE TABLE IF NOT EXISTS users
                         (user_id TEXT, username TEXT, name TEXT, kurs INTEGER, 
                         groupp TEXT, st_or_teach INTEGER, teacher TEXT)""")

    db.commit()


async def create_profile_student(message: types.CallbackQuery, kurs, group):
    cur.execute("SELECT 1 FROM users WHERE user_id == '{}'"
                      .format(message.from_user.id))

    user = cur.fetchone()

    if not user:
        cur.execute("INSERT INTO users VALUES(?, ?, ?, ?, ?, ?, ?)",
                    (message.from_user.id, message.from_user.username, message.from_user.full_name, kurs, group, 0, ''))
        db.commit()


async def create_profile_teacher(message: types.CallbackQuery, teacher):
    cur.execute("SELECT 1 FROM users WHERE user_id == '{}'"
                      .format(message.from_user.id))
    user = cur.fetchone()

    if not user:
        cur.execute("INSERT INTO users VALUES(?, ?, ?, ?, ?, ?, ?)",
                    (message.from_user.id, message.from_user.username, message.from_user.full_name, 0, 0, 1, teacher))
        db.commit()


async def update_profile(callback_query: types.CallbackQuery, user: list):
    cur.execute(
        "UPDATE users SET username = '{}', name = '{}', kurs = '{}', groupp = '{}', "
        "st_or_teach = '{}', teacher = '{}' WHERE user_id = '{}'"
        .format(callback_query.from_user.username, callback_query.from_user.full_name, user[0], user[1], user[2],
                user[3], callback_query.from_user.id))
    db.commit()


async def delete_profile(user_id):
    cur.execute("DELETE FROM users WHERE user_id = '{}'".format(user_id))
    db.commit()


def add_rasp(raw_table: list, kurs):
    cur.execute("INSERT INTO '{}' VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)".format(kurs),
                    (raw_table[0], raw_table[1], raw_table[2], raw_table[3], raw_table[4], raw_table[5], raw_table[6], raw_table[7], raw_table[8], raw_table[9]))
    db.commit()


async def get_rasp(day, kurs, group: str):
    split_group = group.split()
    specialty = split_group[0]
    group_cut = split_group[-1]
    print(group)
    if specialty == group:
        cur.execute("SELECT * FROM '{}' WHERE day = '{}' AND groupp = '{}'".format(kurs, day, group_cut))
        print(1)

    elif group[-1].isdigit() or group[-1] == ')':
        cur.execute("SELECT * FROM '{}' WHERE day = '{}' AND groupp = '{}'".format(kurs, day, group))
        print(2)
    else:
        cur.execute("SELECT * FROM '{}' WHERE day = '{}' AND groupp = '{}' AND specialty = '{}'".format(kurs, day, group_cut, specialty))
        print(3)
    day_rasp = cur.fetchall()
    return day_rasp


async def get_teachers():
    kurses = await get_kurs()
    teachers = list()
    for kurs in kurses:
        if kurs[1] != 'users' and not('old' in kurs[1]):
            cur.execute("SELECT DISTINCT teach1 FROM '{}' WHERE teach1 IS NOT NULL".format(kurs[1]))
            teachers.extend(cur.fetchall())
            cur.execute("SELECT DISTINCT teach3 FROM '{}' WHERE teach3 IS NOT NULL".format(kurs[1]))
            teachers.extend(cur.fetchall())
    unic_teach = list(set(teachers))
    return unic_teach


async def get_user(user_id):
    cur.execute("SELECT kurs, groupp, st_or_teach, teacher FROM users WHERE user_id = '{}'".format(user_id))
    user = cur.fetchone()
    return user


def chisl_or_znam(now):
    delta = now - day_chisl
    if (delta.days // 7) % 2 == 0:
        return 4
    else:
        return 7


# def cell_value(sheet, raw, col):
#     cell = sheet[raw][col]
#     coord = cell.coordinate
#     if not isinstance(cell, MergedCell):
#         return cell.value
#
#     # "Oh no, the cell is merged!"
#     for range in sheet.merged_cells.ranges:
#         if coord in range:
#             return range.start_cell.value
#
#     raise AssertionError('Merged cell is not in any merge range!')


def cell_value(sheet, raw, col):
    cell = sheet[raw][col]
    coord = cell.coordinate
    if not isinstance(cell, MergedCell):
        return cell.value

    # "Oh no, the cell is merged!"
    for range in sheet.merged_cells.ranges:
        if coord in range:
            return range.start_cell.value

    raise AssertionError('Merged cell is not in any merge range!')


async def get_kurs():
    cur.execute("SELECT * FROM sqlite_master WHERE type = 'table'")
    return cur.fetchall()


async def get_groups(kurs):
    cur.execute("SELECT DISTINCT groupp, specialty FROM '{}'".format(kurs))
    groups = cur.fetchall()
    db.commit()
    return groups


def normalize(target_str):
    if target_str is None:
        target_str = ''
    # получаем заменяемое: подставляемое из словаря в цикле
    for i, j in sokr.items():
        # меняем все target_str на подставляемое
        target_str = target_str.strip()
        target_str = (" ".join(target_str.split())).replace(i, j)
    return target_str


# def sort_by_time(rasp: list):
#     rasp = sorted(rasp, key=lambda para: datetime.strftime(para[1].split('-')[0], '%H.%M'))


async def rename_tables(kurs):
    cur.execute("ALTER TABLE '{}' RENAME TO '{}'".format(kurs[1], kurs[1] + '_old'))
    db.commit()


async def get_diff_rasp(day, kurs, group):
    rasp = await get_rasp(day, kurs, group)
    old_rasp = await get_rasp(day, kurs + '_old', group)

    text = f'<b><u>{number_in_days.get(day)}</u></b>\n'


    if len(rasp) < len(old_rasp):
        min_len = len(rasp)
    else:
        min_len = len(old_rasp)

    for i in range(min_len):
        if rasp[i] != old_rasp[i]:
            text += f'> <b><i>{rasp[i][1]}</i>:</b>\n'
            text += f'Числ: {normalize(rasp[i][4])} - {normalize(rasp[i][5])} [{normalize(rasp[i][6])}]\n'
            text += f'Знам: {normalize(rasp[i][7])} - {normalize(rasp[i][8])} [{normalize(rasp[i][9])}]\n'

    if len(rasp) > min_len:
        for i in range(min_len, len(rasp)):
            text += f'> <b><i>{rasp[i][1]}</i>:</b>\n'
            text += f'Числ: {normalize(rasp[i][4])} - {normalize(rasp[i][5])} [{normalize(rasp[i][6])}]\n'
            text += f'Знам: {normalize(rasp[i][7])} - {normalize(rasp[i][8])} [{normalize(rasp[i][9])}]\n'

    if text == f'<b><u>{number_in_days.get(day)}</u></b>\n':
        return ''
    text += '\n\n'
    return text


async def delete_old():
    tables = await get_kurs()
    for table in tables:
        if 'old' in table[1]:
            cur.execute('DROP TABLE IF EXISTS "{}"'.format(table[1]))
            db.commit()


async def get_all_diff():
    kurses = await get_kurs()

    for kurs in kurses:
        if not ('old' in kurs[1]) and kurs[1] != 'users':
            groups = await get_groups(kurs[1])
            groups.sort()

            for i in range(len(groups) - 1):
                if groups[i][0] == groups[i + 1][0] or groups[i][0] == groups[i - 1][0]:
                    group = f'{groups[i][1]} {groups[i][0]}'
                else:
                    group = groups[i][0]

                text = f'<u>Изменения расписания</u>  {kurs[1]} {group}\n\n'
                for day in range(6):
                    text += await get_diff_rasp(day, kurs[1], group)

                cur.execute("SELECT user_id FROM users WHERE kurs = '{}' AND groupp = '{}'".format(kurs[1], group))
                users = cur.fetchall()

                for user in users:
                    await bot.send_message(user[0], text)
            if len(groups) > 1:
                if groups[-1][0] == groups[-2][0]:
                    group = f'{groups[-1][1]} {groups[-1][0]}'
                else:
                    group = groups[-1][0]
            else:
                group = groups[-1][0]

            text = '<u>Изменения расписания</u>'
            for day in range(6):
                text += '\n\n\n'
                text += await get_diff_rasp(day, kurs[1], group)

            cur.execute("SELECT user_id FROM users WHERE kurs = '{}' AND groupp = '{}'".format(kurs[1], group))
            users = cur.fetchall()

            for user in users:
                await bot.send_message(user[0], text)


async def get_teach_rasp(day, message):
    user = await get_user(message.from_user.id)
    rasp = list()
    kurses = await get_kurs()
    ch_or_zn = chisl_or_znam(day)
    for kurs in kurses:
        if kurs[1] != "users" and not('old' in kurs[1]):
            if ch_or_zn == 4:
                cur.execute("SELECT *, '{}' FROM '{}' WHERE day = '{}' AND teach1 = '{}' ".format(kurs[1], kurs[1],
                                                                                                  day.weekday(),
                                                                                                  user[3]))
            else:
                cur.execute("SELECT *, '{}' FROM '{}' WHERE day = '{}' and teach3 = '{}' ".format(kurs[1], kurs[1],
                                                                                                  day.weekday(),
                                                                                                  user[3]))
            rasp1 = list(cur.fetchall())
            rasp.extend(rasp1)
    rasp = sorted(rasp, key=lambda para: datetime.datetime.strptime(para[1].split('-')[0], '%H.%M'))
    text = text = f'<b><u>{number_in_days.get(day.weekday())}</u></b>'
    text = f'<i>{user[3]}   [{day.strftime("%d.%m.%Y")}]</i>\n' + text
    if ch_or_zn == 4:
        text = text + '(числитель):\n\n'
    else:
        text = text + '(знаменатель):\n\n'
    if len(rasp) == 0:
        text = text + "На сегодня занятий не наблюдается...\nОтличный день для отдыха!"
        return text
    k = 0

    for i in range(0, len(rasp)):
        if k != 0 and rasp[i][1] != rasp[i - 1][1]:
            text = text + f') [{rasp[i - 1][ch_or_zn + 2]}]\n\n'
            k = 0
        if rasp[i][ch_or_zn] is not None:
            if rasp[i][1] != rasp[i - 1][1] or i == 0:
                text = text + f'> <b><i>{rasp[i][1]}</i>:</b>\n'
                text = text + f'{normalize(rasp[i][ch_or_zn])} ({rasp[i][10]}: {rasp[i][3]}'
                k += 1
            else:
                text = text + f', {rasp[i][2]}'
    if k != 0:
        text = text + f') [{rasp[len(rasp) - 1][ch_or_zn + 2]}]\n\n'
        k = 0
    return text


async def day_rasp(message: types.Message, day):
    user = await get_user(message.from_user.id)
    ch_or_zn = chisl_or_znam(day)   # узнаем какая сейчас неделя: числитель или знаменатель
    rasp = list(await get_rasp(day.weekday(), user[0], user[1]))  # получаем данные из базы данных, где нужный день недели, курс и группа
    text = f'<b><u>{number_in_days.get(day.weekday())}</u></b>'
    text = f'<i>{user[0]} {user[1]}   [{day.strftime("%d.%m.%Y")}]</i>\n' + text
    if ch_or_zn == 4:
        text = text + '(числитель):\n\n'
    else:
        text = text + '(знаменатель):\n\n'
    j = 1

    for l in range(1, len(rasp)):
        if rasp[0] == rasp[l]:
            j += 1

    for m in range(len(rasp) // j):
        for k in range(j):
            i = m + (len(rasp) // j) * k
            if (rasp[i][ch_or_zn] is not None and rasp[i][ch_or_zn] != rasp[m + (len(rasp) // j) * (k - 1)][
                ch_or_zn]) or (rasp[i][ch_or_zn] is not None and k == 0):
                text = text + f'> <b><i>{rasp[i][1]}</i>:</b>\n'

                if len(normalize(rasp[i][ch_or_zn + 1])) + len(normalize(rasp[i][ch_or_zn + 2])) > 0:
                    if len(normalize(rasp[i][ch_or_zn + 1])) + len(normalize(rasp[i][ch_or_zn + 2])) < 31 and len(
                            normalize(rasp[i][ch_or_zn])) \
                            + len(normalize(rasp[i][ch_or_zn + 1])) + len(normalize(rasp[i][ch_or_zn + 2])) > 31:
                        text = text + f'{normalize(rasp[i][ch_or_zn])}\n{normalize(rasp[i][ch_or_zn + 1])} [{normalize(rasp[i][ch_or_zn + 2])}]'

                    elif (len(normalize(rasp[i][ch_or_zn])) + len(normalize(rasp[i][ch_or_zn + 1])) + len(
                            normalize(rasp[i][ch_or_zn + 2])) < 27) \
                            or (
                            len((normalize(rasp[i][ch_or_zn]))) < 20 and len(normalize(rasp[i][ch_or_zn + 1])) + len(
                        normalize(rasp[i][ch_or_zn + 2])) < 19):
                        if len((normalize(rasp[i][ch_or_zn]))) > 13:
                            text = text + f'{normalize(rasp[i][ch_or_zn])[:13]}... {normalize(rasp[i][ch_or_zn + 1])} [{normalize(rasp[i][ch_or_zn + 2])}]'
                        else:
                            text = text + f'{normalize(rasp[i][ch_or_zn])[:13]} {normalize(rasp[i][ch_or_zn + 1])} [{normalize(rasp[i][ch_or_zn + 2])}]'
                    elif len(normalize(rasp[i][ch_or_zn + 2])) and len(normalize(rasp[i][ch_or_zn])) + len(
                            normalize(rasp[i][ch_or_zn + 1])) < 31:
                        text = text + f'{normalize(rasp[i][ch_or_zn])} {normalize(rasp[i][ch_or_zn + 1])}\n[{normalize(rasp[i][ch_or_zn + 2])}]'

                    elif len(normalize(rasp[i][ch_or_zn])) + len(normalize(rasp[i][ch_or_zn + 1])) > 31 and len(
                            normalize(rasp[i][ch_or_zn + 1])) + len(normalize(rasp[i][ch_or_zn + 2])) > 31:
                        text = text + f'{normalize(rasp[i][ch_or_zn])}\n{normalize(rasp[i][ch_or_zn + 1])}\n[{normalize(rasp[i][ch_or_zn + 2])}]'

                    else:
                        text = text + f'{normalize(rasp[i][ch_or_zn])} {normalize(rasp[i][ch_or_zn + 1])} [{normalize(rasp[i][ch_or_zn + 2])}]'
                else:
                    text = text + f'{normalize(rasp[i][ch_or_zn])}'
                if k > 0:
                    text = text + f' ({k + 1})\n\n'
                else:
                    text = text + '\n\n'

    if text[-3] == ':':
        text += "На сегодня занятий не наблюдается...\nОтличный день для отдыха!"

    return text




def cut_teach(s: str):
    split_s = list()
    if s is not None:
        list_s = s.split()
        j = len(list_s)
        for i in range(len(list_s)):
            if len(list_s[i]) == 4:
                if list_s[i][0].isalpha() and list_s[i][1]== '.' and list_s[i][2].isalpha() and list_s[i][3]== '.':
                    j = i
                    break
        if j == len(list_s):
            split_s.append(" ".join(list_s))
            split_s.append(None)
            split_s.append(None)
        else:
            split_s.append(" ".join(list_s[:j - 1]))
            split_s.append(" ".join(list_s[j - 1: j + 1]))
            split_s.append(" ".join(list_s[j+1:]))

        if split_s[1] is None and split_s[0] is not None:
            for i in range(len(split_s[0])-3):
                if (split_s[0][i].isalpha() and split_s[0][i].istitle()) and split_s[0][i+1] == '.' \
                        and (split_s[0][i+2].isalpha() and split_s[0][i+2].istitle()) \
                        and (split_s[0][i+3] =='.' or split_s[0][i+3].isspace()):
                    j = i

                    spaces = 0
                    while spaces != 2 and j != 0:
                        if split_s[0][j].isspace():
                            j += -1
                            spaces += 1
                        else:
                            j += -1

                    split_s[1] = split_s[0][j+1: i+4].strip()
                    split_s[2] = split_s[0][i+4:].strip()
                    split_s[0] = split_s[0][:j+1].strip()
                    if split_s[1][-1] != '.':
                        split_s[1] = split_s[1] + '.'

                    break
        if split_s[1] is not None:
            for i in range(len(split_s[1])):
                if split_s[1][i:].istitle():
                    split_s[0] = split_s[0] + split_s[1][:i]
                    split_s[1] = split_s[1][i:]
                    break
    else:
        split_s.append(None)
        split_s.append(None)
        split_s.append(None)
    return split_s


def parse_kurs(col_begin):
    book = openpyxl.open("xl.xlsx")  # открытие excel файла
    sheet = book.active              # делаем лист активным

    corr = 0

    if cell_value(sheet, 3, col_begin) == 'Дни недели':
        corr += 1
    if cell_value(sheet, 3, col_begin + 1) == 'Часы звонков':
        corr += 1

    col = corr + col_begin

    # проходим все ячейки с расписанием и записываем данные в списки
    while (cell_value(sheet, 4, col)) is not None:
        row = 5
        group_para = list()
        while row < sheet.max_row:
            if cell_value(sheet, row, 1) is not None:
                print(cell_value(sheet, row, 1))
                para = list()
                para.append(cell_value(sheet, row, 0))
                para.append(cell_value(sheet, row, 1))
                para.append(cell_value(sheet, 3, col))
                para.append(cell_value(sheet, 4, col))
                para.extend(cut_teach(cell_value(sheet, row, col)))
                para.extend(cut_teach(cell_value(sheet, row + 1, col)))

                para[0] = days_in_number.get(para[0])

                split_spec = para[2].split()
                para[2] = ''
                for i in range(1, len(split_spec)):
                    if len(split_spec[i]) > 1:
                        para[2] += split_spec[i][0].upper()

                # for i in range(1, len(para[3])):
                #     if para[3][len(para[3]) - i].isalpha() and para[3][len(para[3]) - i] != '.':
                #         para[3] = para[3][len(para[3]) - i + 2:]
                #         break

                if para[3] != '':
                    if para[3].split()[0] == "Группа" and len(para[3].split()) > 1:
                        para[3] = para[3].split()[1:]
                        para[3] = ''.join(para[3])

                if para[3] == '':
                    para[3] = cell_value(sheet, 4, col)
                    for i in range(len(para[3])):
                        if para[3].isspace():
                            para[3] = para[3][:i]
                            break

                if len(para[3]) > 10:
                    para[3] = (para[3].split())[0]
                print(para[3])
                group_para.append(para)

                row += 2
            else:
                row += 1

        create_table_rasp(cell_value(sheet, 2, col))
        for par in group_para:
            add_rasp(par, cell_value(sheet, 2, col))

        col += 1
        if col == sheet.max_column:
            break


async def parse_xl():
    book = openpyxl.open("xl.xlsx")
    sheet = book.active
    print(sheet.max_row)
    print(sheet.max_column)
    col_begin = 0

    parse_kurs(col_begin)
    for j in range(sheet.max_column):
        if (cell_value(sheet, 4, j)) is None:
            if cell_value(sheet, 4, j+1) is None and cell_value(sheet, 4, j+2) is None:
                return 0

            j += 1
            col_begin = j
            parse_kurs(col_begin)


